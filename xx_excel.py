# -*- encoding: utf-8 -*-
"""
@File    :   xx_excel.py    
@Time    :   2020-06-13 23:18

基于 openpyxl 模块的，Excel自动化操作

"""

# import openpyxl
# from openpyxl.styles import Font
# from openpyxl.styles import PatternFill
# from openpyxl.styles import Border
# from openpyxl.styles import Alignment
# from openpyxl.styles import Side
# from openpyxl.utils import get_column_letter, column_index_from_string
#
# # 01excel基本操作========================================================
#
# fName = ""
# if fName:
#     # 打开Excel文件
#     wb = openpyxl.load_workbook(fName)
# else:
#     # 创建一个新的Excel文件
#     wb = openpyxl.Workbook()
#
# # 创建工作表
# wb.create_sheet(index=0, title="新建工作表")
#
# # 删除工作表
# del (wb['新建工作表'])
#
# # 获取表格的工作表名称列表
# names = wb.sheetnames
# print("sheetnames:%s" % names)
#
# # 获取指定工作表
# sheet = wb["Sheet"]
#
# # 修改工作表标题
# sheet.title = "工作表1"
#
# # 指定单元格方式1
# sheet["A1"] = ""
#
# # 指定单元格方式2
# sheet.cell(row=2, column=2).value = ''
#
# # 获取表格大小
# max_row = sheet.max_row
# max_column = sheet.max_column
# print("ROW:%s COL:%s" % (max_row, max_column))
#
# # 获取单元格的值
# print("A1:%s B2:%s" % (sheet["A1"].value, sheet["B2"].value))
#
# # 获取矩形区域单元格
# sheets = sheet["A1":"C3"]
# for i in sheets:
#     print(i)
#
# # 01=============================================================
#
# # 02 单元格字体风格更改=============================================================
#
# # 字体
# ft = Font(
#     name=u'微软雅黑',
#     size=11,
#     bold=True,
#     # italic=True,  # 斜体
#     # vertAlign='baseline',  # 上下标'subscript','baseline'='none,'superscript'
#     # underline='single',  # 下划线'singleAccounting', 'double', 'single', 'doubleAccounting'
#     # strike=False,  # 删除线
#     color='2F2F4F'
# )
#
# fill = PatternFill(
#     fill_type="solid",
#     start_color='BBBBBB',  # 单元格填充色
#     end_color='BBBBBB'
# )
#
# # 边框   可以选择的值为：'hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed', 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', 'thick', 'thin']
# bd = Border(
#     # border_style:thin;double;hair;dashed
#     left=Side(border_style="thin", color='BBBBBB'),
#     right=Side(border_style="thin", color='BBBBBB'),
#     top=Side(border_style="thin", color='BBBBBB'),
#     bottom=Side(border_style="thin", color='BBBBBB'),
#     diagonal=Side(border_style='thin', color='BBBBBB'),  # 对角线
#     diagonal_direction=1,
#     outline=Side(border_style='slantDashDot', color='BC1717'),  # 外边框
#     vertical=Side(border_style='medium', color="BBBBBB"),  # 竖直线
#     horizontal=Side(border_style='dotted', color="BBBBBB")  # 水平线
# )
# # 对齐方式
# alignment = Alignment(
#     # 水平：'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
#     horizontal='left',
#     # 垂直：'distributed', 'bottom', 'top', 'center', 'justify'
#     vertical='center',
#     text_rotation=0,  # 旋转角度0~180
#     wrap_text=False,  # 文字换行
#     shrink_to_fit=True,  # 自适应宽度，改变文字大小,上一项false
#     indent=0
# )
#
# number_format = 'General'
#
# # protection = Protection(locked=True,hidden=False)
#
#
# # 合并单元格
# sheet.merge_cells('B1:I1')
# sheet["B1"] = '九九乘法表'
# sheet.row_dimensions[1].height = 32.0
# sheet["B1"].font = Font(
#     name=u'微软雅黑',
#     size=18,
#     bold=True,
#     color='2F2F4F'
# )
# sheet["B1"].alignment = Alignment(horizontal='center',vertical='center')
#
# for x in range(2, 10):
#     # 设置行高
#     """
#     行的高度可以设置为0到409之间的整数或浮点值。
#     这个值表示高度的点数。一点等于1/72英寸。默认的行高是12.75。
#     列宽可以设置为0到255之间的整数或浮点数。
#     这个值表示使用默认字体大小时（11点），单元格可以显示的字符数。
#     默认的列宽是8.43个字符。
#     列宽为零或行高为零，将使单元格隐藏。
#     """
#     sheet.row_dimensions[x].height = 18.0
#     for y in range(2, 10):
#         # 设置列宽
#         sheet.column_dimensions[get_column_letter(y)].width = 12.0
#         c = sheet.cell(row=x, column=y)
#         c.font = ft
#         c.fill = fill
#         c.border = bd
#         c.alignment = alignment
#         c.number_format = number_format
#         c.value = "%sX%s=%s" % (x, y, x * y)
#
# for x in range(11, 20):
#     sheet.row_dimensions[x].height = 18.0
#     for y in range(2, 10):
#         # 设置列宽
#         sheet.column_dimensions[get_column_letter(y)].width = 12.0
#         c = sheet.cell(row=x, column=y)
#         c.font = ft
#         c.fill = fill
#         c.border = bd
#         c.alignment = alignment
#         c.number_format = number_format
#         c.value = "%s" % (x * y)
#

from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

wb = Workbook()
ws = wb.active

rows = [
    [0, 1, 2],
    [1, 10, 20],
    [2, 30, 40]
]

for row in rows:
    ws.append(row)

chart = ScatterChart()
chart.title = "Scatter Chart"
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'

xvalues = Reference(ws, min_col=2, min_row=1, max_col=3)
yvalues = Reference(ws, min_col=1, min_row=2, max_row=3)
values = Reference(ws, min_col=2, min_row=2, max_col=3,max_row=3)
series = Series( xvalues,yvalues, title_from_data=True)
chart.series.append(series)

ws.add_chart(chart, "A10")

wb.save("scatter.xlsx")




# 保存表格
wb.save("table9.xlsx")
